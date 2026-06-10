import io
import csv
import logging
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill

from commons.db import get_conversations_collection, get_archived_proxy_collection, get_proxy_collection, get_validation_collection

log = logging.getLogger("export_utils")

def flatten_conversations():
    """
    Fetch all conversations from MongoDB and flatten them into a Pandas DataFrame.
    """
    collection = get_conversations_collection()
    if collection is None:
        log.error("Failed to get conversations collection from DB.")
        return None

    records = []
    
    # iterate through all documents in the collection
    for doc in collection.find({}):
        # Base information
        record = {
            "user_id": doc.get("user_id"),
            "user_name": doc.get("user_name"),
            "cornell_id": doc.get("cornell_id"),
            "personality": doc.get("personality"),
        }
        
        # spc_raw nested object
        spc_raw = doc.get("spc_raw", {})
        
        if isinstance(spc_raw, dict):
            # Personality scores -> p_
            p_scores = spc_raw.get("personality_scores", {})
            if isinstance(p_scores, dict):
                for k, v in p_scores.items():
                    record[f"p_{k}"] = v
                    
            # Value scores -> v_
            v_scores = spc_raw.get("value_scores", {})
            if isinstance(v_scores, dict):
                for k, v in v_scores.items():
                    record[f"v_{k}"] = v
                    
            # Context -> ctx_
            ctx = spc_raw.get("context", {})
            if isinstance(ctx, dict):
                for k, v in ctx.items():
                    record[f"ctx_{k}"] = v

        messages = doc.get("messages", [])
        transcript = []
        if isinstance(messages, list):
            system_skipped = False
            for msg in messages:
                if isinstance(msg, dict):
                    role = msg.get("role", "")
                    content = msg.get("content", "")
                    # Skip the first system prompt
                    if role == "system" and not system_skipped:
                        system_skipped = True
                        continue
                    transcript.append(f"[{role}]: {content}")
                    
        record["full_transcript"] = "\n\n".join(transcript)
        records.append(record)

    df = pd.DataFrame(records)
    return df

def flatten_archived_conversations():
    """
    Fetch all archived proxy conversations from MongoDB and flatten them.
    """
    collection = get_archived_proxy_collection()
    if collection is None:
        log.error("Failed to get archived proxy collection from DB.")
        return None

    records = []
    
    for doc in collection.find({}):
        record = {
            "user_name": doc.get("user_name"),
            "user_id": doc.get("user_id"),
        }
        
        # oa_messages
        oa_messages = doc.get("oa_messages", [])
        oa_transcript = []
        if isinstance(oa_messages, list):
            for msg in oa_messages:
                if isinstance(msg, dict):
                    role = msg.get("role", "")
                    content = msg.get("content", "")
                    oa_transcript.append(f"[{role}]: {content}")
        record["oa_messages"] = "\n\n".join(oa_transcript)
        
        # conversation
        conversation = doc.get("conversation", [])
        conv_transcript = []
        if isinstance(conversation, list):
            for msg in conversation:
                if isinstance(msg, dict):
                    role = msg.get("role", "")
                    content = msg.get("content", "")
                    conv_transcript.append(f"[{role}]: {content}")
        record["conversation"] = "\n\n".join(conv_transcript)
        
        records.append(record)
        
    df = pd.DataFrame(records)
    return df

def flatten_partner_proxy_conversations():
    """
    Fetch all partner proxy conversations from MongoDB and flatten them.
    """
    collection = get_proxy_collection()
    if collection is None:
        log.error("Failed to get proxy collection from DB.")
        return None

    records = []
    
    for doc in collection.find({}):
        record = {
            "user_name": doc.get("user_name"),
            "user_id": doc.get("user_id"),
        }
        
        # conversation
        conversation = doc.get("conversation", [])
        conv_transcript = []
        if isinstance(conversation, list):
            for msg in conversation:
                if isinstance(msg, dict):
                    role = msg.get("role", "")
                    content = msg.get("content", "")
                    
                    if role == "user":
                        conv_transcript.append(f"[user]: {content}")
                    elif role == "assistant":
                        metadata = msg.get("metadata")
                        if not isinstance(metadata, dict):
                            metadata = {}
                            
                        category = metadata.get("category", "")
                        action = metadata.get("action", "")
                        has_prior_knowledge = metadata.get("has_prior_knowledge", "")
                        confidence = metadata.get("confidence", "")
                        extracted_question = metadata.get("extracted_question", "")
                        user_query = metadata.get("user_query", "")
                        
                        # Only append commas if there is actually some metadata
                        if category or action or str(has_prior_knowledge) or confidence or extracted_question or user_query:
                            meta_parts = []
                            if category: meta_parts.append(f"category: {category}")
                            if action: meta_parts.append(f"action: {action}")
                            if has_prior_knowledge is not None and has_prior_knowledge != "": meta_parts.append(f"has_prior_knowledge: {has_prior_knowledge}")
                            if confidence: meta_parts.append(f"confidence: {confidence}")
                            if extracted_question: meta_parts.append(f"extracted_question: {extracted_question}")
                            if user_query: meta_parts.append(f"user_query: {user_query}")
                            
                            meta_str = ", ".join(meta_parts)
                            conv_transcript.append(f"[assistant] (Meta: {meta_str}): {content}")
                        else:
                            conv_transcript.append(f"[assistant]: {content}")
                    else:
                        conv_transcript.append(f"[{role}]: {content}")
                        
        record["conversation"] = "\n\n".join(conv_transcript)
        records.append(record)
        
    df = pd.DataFrame(records)
    return df

def flatten_validation_data():
    """
    Fetch all validation documents from MongoDB and flatten them into a long format.
    """
    collection = get_validation_collection()
    if collection is None:
        log.error("Failed to get validation collection from DB.")
        return None

    records = []
    
    for doc in collection.find({}):
        base_record = {
            "user_id": doc.get("user_id"),
            "average_masked_similarity": doc.get("average_masked_similarity"),
            "average_similarity": doc.get("average_similarity"),
        }
        
        # Flatten masked_conversations (using Option A - Long format)
        masked_convs = doc.get("masked_conversations", [])
        if isinstance(masked_convs, list):
            for turn in masked_convs:
                if isinstance(turn, dict):
                    record = base_record.copy()
                    record["conversation_type"] = "masked"
                    record["question"] = turn.get("question")
                    record["raw_question"] = turn.get("raw_question")
                    record["user_answer"] = turn.get("user_answer")
                    record["proxy_answer"] = turn.get("proxy_answer")
                    record["is_masked"] = turn.get("is_masked")
                    record["similarity"] = turn.get("similarity")
                    records.append(record)

        # Flatten conversation
        convs = doc.get("conversation", [])
        if isinstance(convs, list):
            for turn in convs:
                if isinstance(turn, dict):
                    record = base_record.copy()
                    record["conversation_type"] = "unmasked"
                    record["question"] = turn.get("question")
                    record["raw_question"] = turn.get("raw_question")
                    record["user_answer"] = turn.get("user_answer")
                    record["proxy_answer"] = turn.get("proxy_answer")
                    record["is_masked"] = turn.get("is_masked")
                    record["similarity"] = turn.get("similarity")
                    records.append(record)
                    
    df = pd.DataFrame(records)
    return df

def export_conversations_to_excel_and_upload(web_client, channel_id):
    """
    Fetch all data, convert to a single Excel file with multiple tabs, and upload to Slack.
    """
    log.info(f"Starting Excel export for channel {channel_id}")
    try:
        timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"exported_data_{timestamp_str}.xlsx"
        
        # Create an in-memory buffer for the Excel file
        excel_buffer = io.BytesIO()
        
        # Use pandas ExcelWriter
        has_data = False
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # 1. Conversations
            df = flatten_conversations()
            if df is not None and not df.empty:
                df.to_excel(writer, sheet_name="Conversations", index=False)
                has_data = True
                
            # 2. Archived Proxy Conversations
            df_archived = flatten_archived_conversations()
            if df_archived is not None and not df_archived.empty:
                df_archived.to_excel(writer, sheet_name="Archived Proxy", index=False)
                has_data = True
                
            # 3. Partner Proxy Conversations
            df_partner = flatten_partner_proxy_conversations()
            if df_partner is not None and not df_partner.empty:
                df_partner.to_excel(writer, sheet_name="Partner Proxy", index=False)
                has_data = True
                
            # 4. Validation Results
            df_validation = flatten_validation_data()
            if df_validation is not None and not df_validation.empty:
                # Grouping/cleaning data before export
                # Keep user_id, average_masked_similarity, and average_similarity only on the first row per user
                mask = df_validation.duplicated(subset=['user_id'])
                
                # Convert columns to object type to safely insert empty strings without dtype errors
                cols_to_clear = ['user_id', 'average_masked_similarity', 'average_similarity']
                for col in cols_to_clear:
                    if col in df_validation.columns:
                        df_validation[col] = df_validation[col].astype(object)
                
                df_validation.loc[mask, cols_to_clear] = ""

                df_validation.to_excel(writer, sheet_name="Validation Results", index=False)
                has_data = True
                
                # Format the Validation Results sheet
                worksheet = writer.sheets["Validation Results"]
                worksheet.freeze_panes = "A2"
                
                header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                header_font = Font(bold=True)
                
                for idx, col_name in enumerate(df_validation.columns, 1):
                    # Style the header
                    cell = worksheet.cell(row=1, column=idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    
                    # Number formatting for similarity columns
                    if "similarity" in col_name:
                        for row_idx in range(2, len(df_validation) + 2):
                            val_cell = worksheet.cell(row=row_idx, column=idx)
                            # Apply number format if the cell is not empty
                            if val_cell.value != "":
                                val_cell.number_format = "0.000"
                
        if not has_data:
            log.warning("No data found or failed to fetch for all collections.")
            web_client.chat_postMessage(
                channel=channel_id,
                text="Failed to fetch data or the database is completely empty."
            )
            return

        # Get the value of the buffer
        excel_content = excel_buffer.getvalue()

        # Upload to Slack using the files_upload_v2 API
        log.info(f"Uploading file {filename} to channel {channel_id}")
        response = web_client.files_upload_v2(
            channel=channel_id,
            content=excel_content,
            filename=filename,
            title="Exported Data",
            initial_comment="✅ Export of all conversation data (Excel) is complete."
        )
        log.info(f"File upload successful: {response.get('file', {}).get('id')}")

    except Exception as e:
        log.error(f"Error during export and upload: {e}", exc_info=True)
        try:
            web_client.chat_postMessage(
                channel=channel_id,
                text=f"❌ An error occurred during data export: {str(e)}"
            )
        except Exception as slack_error:
            log.error(f"Failed to send error message to slack: {slack_error}")
